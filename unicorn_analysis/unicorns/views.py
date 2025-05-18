from django.shortcuts import render
from .models import UnicornCompany
from django.db.models import Avg, Count, Sum
from django.db.models.functions import TruncYear
import io
import json

import pandas as pd
from django.http import HttpResponse, JsonResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

def home(request):
    unicorns = UnicornCompany.objects.all()
    return render(request, 'unicorns/home.html', {'unicorns': unicorns})

def analysis(request):
    # Get filter parameters from request
    country_filter = request.GET.get('country', '')
    industry_filter = request.GET.get('industry', '')
    financial_stage_filter = request.GET.get('financial_stage', '')

    # Base queryset with filters applied
    unicorns_qs = UnicornCompany.objects.all()
    if country_filter:
        unicorns_qs = unicorns_qs.filter(country=country_filter)
    if industry_filter:
        unicorns_qs = unicorns_qs.filter(industry=industry_filter)
    if financial_stage_filter:
        unicorns_qs = unicorns_qs.filter(financial_stage=financial_stage_filter)

    total_unicorns = unicorns_qs.count()
    avg_valuation = unicorns_qs.aggregate(Avg('valuation'))['valuation__avg'] or 0

    # Count by country
    country_counts = unicorns_qs.values('country').annotate(count=Count('id')).order_by('-count')
    countries = [entry['country'] for entry in country_counts]
    country_data = [entry['count'] for entry in country_counts]

    # Count by industry
    industry_counts = unicorns_qs.values('industry').annotate(count=Count('id')).order_by('-count')
    industries = [entry['industry'] for entry in industry_counts]
    industry_data = [entry['count'] for entry in industry_counts]

    # Count by financial stage
    financial_stage_counts = unicorns_qs.values('financial_stage').annotate(count=Count('id')).order_by('-count')
    financial_stages = [entry['financial_stage'] or 'Unknown' for entry in financial_stage_counts]
    financial_stage_data = [entry['count'] for entry in financial_stage_counts]

    # Average investors count
    avg_investors_count = unicorns_qs.aggregate(Avg('investors_count'))['investors_count__avg'] or 0

    # Total portfolio exits
    total_portfolio_exits = unicorns_qs.aggregate(total=Count('portfolio_exits'))['total'] or 0

    # Top countries by total valuation
    top_countries_valuation = unicorns_qs.values('country').annotate(total_valuation=Sum('valuation')).order_by('-total_valuation')[:10]
    top_countries = [entry['country'] for entry in top_countries_valuation]
    top_countries_valuation_data = [float(entry['total_valuation'])/1e9 if entry['total_valuation'] else 0 for entry in top_countries_valuation]

    # Top industries by total valuation - filter out invalid industry names containing commas or empty
    valid_industries_qs = unicorns_qs.exclude(industry__icontains=',').exclude(industry__exact='').values('industry')
    top_industries_valuation = valid_industries_qs.annotate(total_valuation=Sum('valuation')).order_by('-total_valuation')[:10]
    top_industries = [entry['industry'] for entry in top_industries_valuation]
    top_industries_valuation_data = [float(entry['total_valuation'])/1e9 if entry['total_valuation'] else 0 for entry in top_industries_valuation]

    # Time series data for unicorns joined per year
    joined_per_year_qs = unicorns_qs.annotate(year=TruncYear('date_joined')).values('year').annotate(count=Count('id')).order_by('year')
    joined_years = [entry['year'].year if entry['year'] else 'Unknown' for entry in joined_per_year_qs]
    joined_counts = [entry['count'] for entry in joined_per_year_qs]

    # Time series data for unicorns founded per year
    founded_per_year_qs = unicorns_qs.values('founded_year').annotate(count=Count('id')).order_by('founded_year')
    founded_years = [entry['founded_year'] for entry in founded_per_year_qs if entry['founded_year'] > 0]
    founded_counts = [entry['count'] for entry in founded_per_year_qs if entry['founded_year'] > 0]

    # Data for detailed table
    unicorns_table = unicorns_qs.order_by('-valuation').values(
        'name', 'valuation', 'date_joined', 'country', 'city', 'industry', 'investors',
        'founded_year', 'total_raised', 'financial_stage', 'investors_count', 'deal_terms', 'portfolio_exits'
    )

    # Get distinct filter options for UI
    all_countries = UnicornCompany.objects.values_list('country', flat=True).distinct().order_by('country')
    all_industries = UnicornCompany.objects.values_list('industry', flat=True).distinct().order_by('industry')
    all_financial_stages = UnicornCompany.objects.values_list('financial_stage', flat=True).distinct().order_by('financial_stage')

    context = {
        'total_unicorns': total_unicorns,
        'avg_valuation': round(avg_valuation, 2),
        'countries': countries,
        'country_data': country_data,
        'industries': industries,
        'industry_data': industry_data,
        'financial_stages': financial_stages,
        'financial_stage_data': financial_stage_data,
        'avg_investors_count': round(avg_investors_count, 2),
        'total_portfolio_exits': total_portfolio_exits,
        'top_countries': top_countries,
        'top_countries_valuation_data': top_countries_valuation_data,
        'top_industries': top_industries,
        'top_industries_valuation_data': top_industries_valuation_data,
        'joined_years': json.dumps(joined_years),
        'joined_counts': json.dumps(joined_counts),
        'founded_years': json.dumps(founded_years),
        'founded_counts': json.dumps(founded_counts),
        'unicorns_table': unicorns_table,
        'all_countries': all_countries,
        'all_industries': all_industries,
        'all_financial_stages': all_financial_stages,
        'selected_country': country_filter,
        'selected_industry': industry_filter,
        'selected_financial_stage': financial_stage_filter,
    }
    return render(request, 'unicorns/analysis.html', context)

def export_excel(request):
    # Export filtered unicorns data as Excel with charts recreated
    country_filter = request.GET.get('country', '')
    industry_filter = request.GET.get('industry', '')
    financial_stage_filter = request.GET.get('financial_stage', '')

    unicorns_qs = UnicornCompany.objects.all()
    if country_filter:
        unicorns_qs = unicorns_qs.filter(country=country_filter)
    if industry_filter:
        unicorns_qs = unicorns_qs.filter(industry=industry_filter)
    if financial_stage_filter:
        unicorns_qs = unicorns_qs.filter(financial_stage=financial_stage_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Unicorns Data"

    headers = ['Name', 'Valuation', 'Date Joined', 'Country', 'City', 'Industry', 'Investors',
               'Founded Year', 'Total Raised', 'Financial Stage', 'Investors Count', 'Deal Terms', 'Portfolio Exits']
    ws.append(headers)

    for unicorn in unicorns_qs:
        ws.append([
            unicorn.name,
            unicorn.valuation,
            unicorn.date_joined.strftime('%Y-%m-%d') if unicorn.date_joined else '',
            unicorn.country,
            unicorn.city,
            unicorn.industry,
            unicorn.investors,
            unicorn.founded_year,
            unicorn.total_raised,
            unicorn.financial_stage,
            unicorn.investors_count,
            unicorn.deal_terms,
            unicorn.portfolio_exits
        ])

    # Prepare data for charts
    country_counts = unicorns_qs.values('country').annotate(count=Count('id')).order_by('-count')
    countries = [entry['country'] for entry in country_counts]
    country_data = [entry['count'] for entry in country_counts]

    industry_counts = unicorns_qs.values('industry').annotate(count=Count('id')).order_by('-count')
    industries = [entry['industry'] for entry in industry_counts]
    industry_data = [entry['count'] for entry in industry_counts]

    financial_stage_counts = unicorns_qs.values('financial_stage').annotate(count=Count('id')).order_by('-count')
    financial_stages = [entry['financial_stage'] or 'Unknown' for entry in financial_stage_counts]
    financial_stage_data = [entry['count'] for entry in financial_stage_counts]

    # Add charts to Excel
    def add_pie_chart(ws, data_labels, data_values, title, pos):
        from openpyxl.chart import PieChart
        chart = PieChart()
        chart.title = title
        data_ref = Reference(ws, min_col=pos+2, min_row=2, max_row=len(data_values)+1)
        labels_ref = Reference(ws, min_col=pos+1, min_row=2, max_row=len(data_labels)+1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, f"{chr(65+pos)}15")

    # Write chart data to sheet
    start_row = 2
    ws.cell(row=1, column=15, value="Country")
    ws.cell(row=1, column=16, value="Count")
    for i, (label, value) in enumerate(zip(countries, country_data), start=start_row):
        ws.cell(row=i, column=15, value=label)
        ws.cell(row=i, column=16, value=value)

    ws.cell(row=1, column=18, value="Industry")
    ws.cell(row=1, column=19, value="Count")
    for i, (label, value) in enumerate(zip(industries, industry_data), start=start_row):
        ws.cell(row=i, column=18, value=label)
        ws.cell(row=i, column=19, value=value)

    ws.cell(row=1, column=21, value="Financial Stage")
    ws.cell(row=1, column=22, value="Count")
    for i, (label, value) in enumerate(zip(financial_stages, financial_stage_data), start=start_row):
        ws.cell(row=i, column=21, value=label)
        ws.cell(row=i, column=22, value=value)

    # Add pie charts
    add_pie_chart(ws, countries, country_data, "Distribution by Country", 14)
    add_pie_chart(ws, industries, industry_data, "Distribution by Industry", 17)
    add_pie_chart(ws, financial_stages, financial_stage_data, "Distribution by Financial Stage", 20)

    # Save to BytesIO and return response
    with io.BytesIO() as b:
        wb.save(b)
        b.seek(0)
        response = HttpResponse(
            b.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="unicorns.xlsx"'
        return response

# Removed map_data view as Geographic Map is removed from UI

def detailed_unicorns(request):
    unicorns = UnicornCompany.objects.all()
    return render(request, 'unicorns/detailed_unicorns.html', {'unicorns': unicorns})

